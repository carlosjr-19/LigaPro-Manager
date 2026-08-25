from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from models import Match, League, Court, Team, OwnerCourtSetting, IgnoredDiscrepancy, ArchivedFinance
from extensions import db
from sqlalchemy import func
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
import hashlib

def process_archives(archived_finances, group_by, financial_data, court_totals, total_month_profit=None):
    for archive in archived_finances:
        match_date = datetime.combine(archive.date, datetime.min.time())
        if group_by == 'week':
            start_of_week = match_date - timedelta(days=match_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            date_key = f"Semana {start_of_week.isocalendar()[1]} ({start_of_week.strftime('%d/%m')} - {end_of_week.strftime('%d/%m')})"
            date_obj_sort = start_of_week
        elif group_by == 'month':
            months_es_dict = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
            month_name = months_es_dict.get(match_date.month, "")
            date_key = f"{month_name} {match_date.year}"
            date_obj_sort = match_date.replace(day=1)
        else: # day
            date_key = match_date.strftime('%d/%m/%Y')
            date_obj_sort = match_date

        court_name = archive.court_name
        income = archive.income
        expense = archive.expense
        profit = archive.profit
        
        if date_key not in financial_data:
            financial_data[date_key] = {
                'date_obj': date_obj_sort,
                'display_date': date_key,
                'courts': {},
                'daily_total': 0,
                'daily_income': 0,
                'daily_expense': 0,
                'daily_profit': 0
            }
            
        if court_name not in financial_data[date_key]['courts']:
            financial_data[date_key]['courts'][court_name] = {'income': 0, 'expense': 0, 'profit': 0}
            
        financial_data[date_key]['courts'][court_name]['income'] += income
        financial_data[date_key]['courts'][court_name]['expense'] += expense
        financial_data[date_key]['courts'][court_name]['profit'] += profit
        
        if 'daily_total' in financial_data[date_key]:
            financial_data[date_key]['daily_total'] += profit
        if 'daily_income' in financial_data[date_key]:
            financial_data[date_key]['daily_income'] += income
        if 'daily_expense' in financial_data[date_key]:
            financial_data[date_key]['daily_expense'] += expense
        if 'daily_profit' in financial_data[date_key]:
            financial_data[date_key]['daily_profit'] += profit
            
        if total_month_profit is not None:
            total_month_profit += profit
            
        if court_name not in court_totals:
            court_totals[court_name] = {'income': 0, 'expense': 0, 'profit': 0, 'dates': {}}
            
        court_totals[court_name]['income'] += income
        court_totals[court_name]['expense'] += expense
        court_totals[court_name]['profit'] += profit
        
        if 'dates' in court_totals[court_name]:
            if date_key not in court_totals[court_name]['dates']:
                court_totals[court_name]['dates'][date_key] = {
                    'date_obj': date_obj_sort,
                    'display_date': date_key,
                    'income': 0,
                    'expense': 0,
                    'profit': 0
                }
            court_totals[court_name]['dates'][date_key]['income'] += income
            court_totals[court_name]['dates'][date_key]['expense'] += expense
            court_totals[court_name]['dates'][date_key]['profit'] += profit

    return total_month_profit
    
class FakeTeam:
    def __init__(self, name):
        self.name = name

class FakeCourt:
    def __init__(self, name):
        self.name = name
        self.color = None

class FakeLeague:
    def __init__(self, name, price):
        self.name = name
        self.id = 0
        self.custom_color_active = False
        self.custom_name_color = None
        self.price_per_match = price
        self.price_referee = 0

class FakeMatch:
    def __init__(self, m_data, league_name, date_str):
        import datetime as dt
        self.id = f"arc_{m_data.get('arc_id')}_{m_data.get('match_idx', 0)}"
        self.home_team = FakeTeam(m_data.get('home', 'Local'))
        self.away_team = FakeTeam(m_data.get('away', 'Visita'))
        self.court = FakeCourt(m_data.get('court_name', 'Sin Cancha'))
        self.league = FakeLeague(league_name, m_data.get('expected_price', 0))
        
        # Determine match_date
        raw_date = m_data.get('match_date_raw')
        if raw_date:
            try:
                self.match_date = dt.datetime.strptime(raw_date, '%Y-%m-%d %H:%M:%S')
            except:
                self.match_date = dt.datetime.strptime(date_str, '%Y-%m-%d')
        else:
            time_str = m_data.get('time', '')
            if time_str:
                try:
                    self.match_date = dt.datetime.strptime(f"{date_str} {time_str}", '%Y-%m-%d %I:%M %p')
                except:
                    try:
                        self.match_date = dt.datetime.strptime(f"{date_str} {time_str}", '%Y-%m-%d %H:%M')
                    except:
                        self.match_date = dt.datetime.strptime(date_str, '%Y-%m-%d')
            else:
                self.match_date = dt.datetime.strptime(date_str, '%Y-%m-%d')
                
        self.referee_cost_home = str(m_data.get('referee_cost_home_raw', m_data.get('home_paid', 0)))
        self.referee_cost_away = str(m_data.get('referee_cost_away_raw', m_data.get('away_paid', 0)))
        self.referee_cost = str(m_data.get('referee_cost_raw', m_data.get('ref_paid', 0)))
        
        self.home_score = m_data.get('home_score')
        self.away_score = m_data.get('away_score')
        self.is_practice = m_data.get('is_practice', False)
        self.is_archived = True
        self.home_team_id = None
        self.away_team_id = None

report_bp = Blueprint('report', __name__)

@report_bp.route('/report')
@login_required
def index():
    if current_user.role not in ['owner', 'admin']:
        flash('No tienes permiso para acceder a esta sección.', 'danger')
        return redirect(url_for('main.captain_dashboard'))
    return render_template('report.html')

@report_bp.route('/global-schedule')
@login_required
def global_schedule():
    # Ultra Premium Check
    if not getattr(current_user, 'is_ultra', False):
        flash('No tienes acceso a esta funcionalidad (Ultra Premium).', 'warning')
        return redirect(url_for('report.index'))

    # Date Parameter
    date_str = request.args.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()

    # Time filter parameters (HH:MM format, 24h)
    time_from_str = request.args.get('time_from', '').strip()
    time_to_str = request.args.get('time_to', '').strip()
    time_from = None
    time_to = None
    try:
        if time_from_str:
            time_from = datetime.strptime(time_from_str, '%H:%M').time()
    except ValueError:
        pass
    try:
        if time_to_str:
            time_to = datetime.strptime(time_to_str, '%H:%M').time()
    except ValueError:
        pass

    # Query Matches for ALL leagues owned by current_user on selected_date
    matches = Match.query.join(League).filter(
        League.user_id == current_user.id,
        func.date(Match.match_date) == selected_date
    ).order_by(Match.match_date).all()
    
    # Inject Archived Matches
    archived = ArchivedFinance.query.filter(
        ArchivedFinance.user_id == current_user.id,
        ArchivedFinance.date == selected_date
    ).all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, selected_date.strftime('%Y-%m-%d'))
                    matches.append(fake_m)
            except:
                pass
                
    # Sort again just in case
    matches.sort(key=lambda x: x.match_date if x.match_date else datetime.max)

    # Apply time filter in Python (compatible with SQLite & PostgreSQL)
    if time_from or time_to:
        filtered = []
        for m in matches:
            t = m.match_date.time()
            if time_from and t < time_from:
                continue
            if time_to and t > time_to:
                continue
            filtered.append(m)
        matches = filtered

    # Group by Court Name (String Select)
    grouped_schedule = {}
    
    # Pre-load owner settings to avoid N+1 queries
    owner_settings = OwnerCourtSetting.query.filter_by(user_id=current_user.id).all()
    owner_colors = {s.court_name: s.color for s in owner_settings}
    
    for match in matches:
        # Determine Court Name
        if match.court:
            court_name = match.court.name
            court_color = owner_colors.get(court_name, match.court.color)
        else:
            court_name = "Sin Cancha Asignada"
            court_color = None
            
        if court_name not in grouped_schedule:
            grouped_schedule[court_name] = {
                'matches': [],
                'total_cost_home': 0,
                'total_cost_away': 0,
                'total_referee': 0,
                'total_profit': 0,
                'court_color': court_color
            }
        
        def safe_int(val):
            try:
                return int(val)
            except (ValueError, TypeError):
                return 0

        grouped_schedule[court_name]['matches'].append(match)
        vis_home = safe_int(match.referee_cost_home)
        vis_away = safe_int(match.referee_cost_away)
        exp_ref = safe_int(match.referee_cost)
        
        grouped_schedule[court_name]['total_cost_home'] += vis_home
        grouped_schedule[court_name]['total_cost_away'] += vis_away
        grouped_schedule[court_name]['total_referee'] += exp_ref
        grouped_schedule[court_name]['total_profit'] += ((vis_home + vis_away) - exp_ref)

    # Detect conflicts (same date+time in same court)
    conflicting_match_ids = set()
    for court_name, data in grouped_schedule.items():
        time_counts = {}
        for match in data['matches']:
            if not match.match_date: continue
            time_counts[match.match_date] = time_counts.get(match.match_date, 0) + 1
            
        for match in data['matches']:
            if match.match_date and time_counts[match.match_date] > 1:
                conflicting_match_ids.add(match.id)

    # Sort groups by name
    sorted_schedule = dict(sorted(grouped_schedule.items()))

    return render_template('report/global_schedule.html', 
                         schedule=sorted_schedule, 
                         selected_date=selected_date,
                         conflicting_match_ids=conflicting_match_ids,
                         time_from=time_from_str,
                         time_to=time_to_str)

@report_bp.route('/global-schedule/share')
@login_required
def share_global_schedule():
    # Ultra Premium Check
    if not getattr(current_user, 'is_ultra', False):
        flash('No tienes acceso a esta funcionalidad (Ultra Premium).', 'warning')
        return redirect(url_for('report.index'))

    # Date Parameter
    date_str = request.args.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()

    # Time filter parameters (HH:MM format, 24h)
    time_from_str = request.args.get('time_from', '').strip()
    time_to_str = request.args.get('time_to', '').strip()
    time_from = None
    time_to = None
    try:
        if time_from_str:
            time_from = datetime.strptime(time_from_str, '%H:%M').time()
    except ValueError:
        pass
    try:
        if time_to_str:
            time_to = datetime.strptime(time_to_str, '%H:%M').time()
    except ValueError:
        pass

    # Query Matches for ALL leagues owned by current_user on selected_date
    matches = Match.query.join(League).filter(
        League.user_id == current_user.id,
        func.date(Match.match_date) == selected_date
    ).order_by(Match.match_date).all()
    
    # Inject Archived Matches
    archived = ArchivedFinance.query.filter(
        ArchivedFinance.user_id == current_user.id,
        ArchivedFinance.date == selected_date
    ).all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, selected_date.strftime('%Y-%m-%d'))
                    matches.append(fake_m)
            except:
                pass
                
    # Sort again just in case
    matches.sort(key=lambda x: x.match_date if x.match_date else datetime.max)

    # Apply time filter in Python
    if time_from or time_to:
        filtered = []
        for m in matches:
            t = m.match_date.time()
            if time_from and t < time_from:
                continue
            if time_to and t > time_to:
                continue
            filtered.append(m)
        matches = filtered

    # Group by Court Name (String Select)
    grouped_schedule = {}
    
    owner_settings = OwnerCourtSetting.query.filter_by(user_id=current_user.id).all()
    owner_colors = {s.court_name: s.color for s in owner_settings}

    for match in matches:
        if match.court:
            court_name = match.court.name
            court_color = owner_colors.get(court_name, match.court.color)
        else:
            court_name = "Sin Cancha Asignada"
            court_color = None
            
        if court_name not in grouped_schedule:
            grouped_schedule[court_name] = {
                'matches': [],
                'total_cost_home': 0,
                'total_cost_away': 0,
                'total_referee': 0,
                'total_profit': 0,
                'court_color': court_color
            }
        
        def safe_int(val):
            try:
                return int(val)
            except (ValueError, TypeError):
                return 0

        grouped_schedule[court_name]['matches'].append(match)
        vis_home = safe_int(match.referee_cost_home)
        vis_away = safe_int(match.referee_cost_away)
        exp_ref = safe_int(match.referee_cost)
        
        grouped_schedule[court_name]['total_cost_home'] += vis_home
        grouped_schedule[court_name]['total_cost_away'] += vis_away
        grouped_schedule[court_name]['total_referee'] += exp_ref
        grouped_schedule[court_name]['total_profit'] += ((vis_home + vis_away) - exp_ref)

    # Detect conflicts
    conflicting_match_ids = set()
    for court_name, data in grouped_schedule.items():
        time_counts = {}
        for match in data['matches']:
            if not match.match_date: continue
            time_counts[match.match_date] = time_counts.get(match.match_date, 0) + 1
            
        for match in data['matches']:
            if match.match_date and time_counts[match.match_date] > 1:
                conflicting_match_ids.add(match.id)

    sorted_schedule = dict(sorted(grouped_schedule.items()))

    # Build teams dict for easy shield retrieval in template
    teams_dict = {}
    for match in matches:
        if match.home_team_id not in teams_dict:
            teams_dict[match.home_team_id] = Team.query.get(match.home_team_id)
        if match.away_team_id not in teams_dict:
            teams_dict[match.away_team_id] = Team.query.get(match.away_team_id)

    return render_template('report/share_global_schedule.html', 
                         schedule=sorted_schedule, 
                         selected_date=selected_date,
                         conflicting_match_ids=conflicting_match_ids,
                         teams_dict=teams_dict,
                         time_from=time_from_str,
                         time_to=time_to_str,
                         today=datetime.now().strftime('%d/%m/%Y'))

@report_bp.route('/api/match/update_costs', methods=['POST'])
@login_required
def update_match_costs():
    if not getattr(current_user, 'is_ultra', False):
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.get_json()
    match_id = data.get('match_id')
    
    # Fields to update
    cost_home = data.get('referee_cost_home')
    cost_away = data.get('referee_cost_away')
    cost_referee = data.get('referee_cost')
    home_score = data.get('home_score')
    away_score = data.get('away_score')
    
    match = Match.query.get(match_id)
    if not match:
        return jsonify({'error': 'Match not found'}), 404
        
    # Update cost fields if provided
    if cost_home is not None:
        match.referee_cost_home = str(cost_home)
    if cost_away is not None:
        match.referee_cost_away = str(cost_away)
    if cost_referee is not None:
        match.referee_cost = str(cost_referee)
    
    # Update scores
    if home_score is not None:
        try:
            match.home_score = int(home_score) if str(home_score).strip() != "" else None
        except ValueError:
            pass
            
    if away_score is not None:
        try:
            match.away_score = int(away_score) if str(away_score).strip() != "" else None
        except ValueError:
            pass

    # Auto-complete if scores are present (mirroring match_matrix logic)
    if match.home_score is not None and match.away_score is not None:
        match.is_completed = True
    else:
        match.is_completed = False
    
    db.session.commit()
    
    return jsonify({'success': True, 'match_id': match.id, 'is_completed': match.is_completed})

@report_bp.route('/global-schedule/config', methods=['GET', 'POST'])
@login_required
def global_schedule_config():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))
        
    if request.method == 'POST':
        leagues = League.query.filter_by(user_id=current_user.id).all()
        for league in leagues:
            # Prices
            price_team = request.form.get(f'price_team_{league.id}')
            price_ref = request.form.get(f'price_referee_{league.id}')
            
            # Start Date Settings
            charge_from_start = request.form.get(f'charge_from_start_{league.id}') == 'on'
            auto_fill_prices = request.form.get(f'auto_fill_prices_{league.id}') == 'on'
            charge_date_str = request.form.get(f'charge_start_date_{league.id}')
            
            try:
                if price_team is not None:
                    league.price_per_match = int(price_team)
                if price_ref is not None:
                    league.price_referee = int(price_ref)
                
                league.charge_from_start = charge_from_start
                league.auto_fill_prices = auto_fill_prices
                league.charge_from_start = charge_from_start
                if not charge_from_start and charge_date_str:
                    league.charge_start_date = datetime.strptime(charge_date_str, '%Y-%m-%d').date()
                else:
                    league.charge_start_date = None
            except ValueError:
                pass
                
        db.session.commit()
        flash('Configuración de precios y fechas actualizada correctamente.', 'success')
        return redirect(url_for('report.global_schedule_config'))

    leagues = League.query.filter_by(user_id=current_user.id).all()
    return render_template('report/config.html', leagues=leagues)

@report_bp.route('/global-schedule/history')
@login_required
def global_schedule_history():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    # Filters
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid]
    
    query = Match.query.join(League).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, [])
        
    matches = query.order_by(Match.match_date.desc()).all()
    archived = archive_query.all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, arc.date.strftime('%Y-%m-%d'))
                    # Set defaults for fake league so it bypasses charge_start_date logic safely
                    fake_m.league.charge_from_start = True
                    fake_m.league.charge_start_date = None
                    matches.append(fake_m)
            except:
                pass
                
    matches.sort(key=lambda x: x.match_date if x.match_date else datetime.max, reverse=True)
    
    show_hidden = request.args.get('show_hidden', '0') == '1'
    from models import IgnoredDiscrepancy
    ignored_records = {x.hash_id for x in IgnoredDiscrepancy.query.filter_by(user_id=current_user.id).all()}
    
    history_events = []
    
    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0 # NSP or text = 0 paid
        try: return int(val)
        except: return 0

    for match in matches:
        if not match.league: continue
        
        # Respect Charge Start Date
        if not match.league.charge_from_start and match.league.charge_start_date:
            if match.match_date.date() < match.league.charge_start_date:
                continue
        
        # Defaults
        default_team_price = match.league.price_per_match or 0
        default_ref_price = match.league.price_referee or 0
        
        # Check Home Team Debt – skip if waived (gifted)
        hash_home = f"{match.id}_home"
        if not is_waived(match.referee_cost_home):
            paid_home = parse_cost(match.referee_cost_home)
            diff_home = paid_home - default_team_price
            if diff_home != 0:
                is_hidden = hash_home in ignored_records
                if show_hidden or not is_hidden:
                    history_events.append({
                        'hash_id': hash_home,
                        'is_hidden': is_hidden,
                        'date': match.match_date,
                        'league': match.league.name,
                        'match': f"{match.home_team.name} vs {match.away_team.name}",
                        'entity': f"Local: {match.home_team.name}",
                        'expected': default_team_price,
                        'paid': paid_home,
                        'balance': diff_home
                    })
            
        # Check Away Team Debt – skip if waived (gifted)
        hash_away = f"{match.id}_away"
        if not is_waived(match.referee_cost_away):
            paid_away = parse_cost(match.referee_cost_away)
            diff_away = paid_away - default_team_price
            if diff_away != 0:
                is_hidden = hash_away in ignored_records
                if show_hidden or not is_hidden:
                    history_events.append({
                        'hash_id': hash_away,
                        'is_hidden': is_hidden,
                        'date': match.match_date,
                        'league': match.league.name,
                        'match': f"{match.home_team.name} vs {match.away_team.name}",
                        'entity': f"Visita: {match.away_team.name}",
                        'expected': default_team_price,
                        'paid': paid_away,
                        'balance': diff_away
                    })

        # Check Referee Balance – skip if waived
        hash_ref = f"{match.id}_ref"
        if not is_waived(match.referee_cost):
            paid_ref = parse_cost(match.referee_cost)
            diff_ref = paid_ref - default_ref_price
            if diff_ref != 0:
                is_hidden = hash_ref in ignored_records
                if show_hidden or not is_hidden:
                    history_events.append({
                        'hash_id': hash_ref,
                        'is_hidden': is_hidden,
                        'date': match.match_date,
                        'league': match.league.name,
                        'match': f"{match.home_team.name} vs {match.away_team.name}",
                        'entity': "Arbitro",
                        'expected': default_ref_price,
                        'paid': paid_ref,
                        'balance': diff_ref
                    })

            
    # Sort events by date desc
    history_events.sort(key=lambda x: x['date'], reverse=True)
    
    leagues = League.query.filter_by(user_id=current_user.id).all()
    
    return render_template('report/history.html', 
                         events=history_events, 
                         leagues=leagues, 
                         selected_league_ids=league_ids,
                         show_hidden=show_hidden)

@report_bp.route('/global-schedule/history/toggle_hide', methods=['POST'])
@login_required
def toggle_hide_history():
    if not getattr(current_user, 'is_ultra', False):
        return jsonify({'success': False, 'message': 'Acceso denegado.'}), 403
        
    hash_id = request.form.get('hash_id')
    if not hash_id:
        return jsonify({'success': False, 'message': 'ID no proporcionado.'}), 400
        
    from models import IgnoredDiscrepancy
    existing = IgnoredDiscrepancy.query.filter_by(user_id=current_user.id, hash_id=hash_id).first()
    
    if existing:
        db.session.delete(existing)
        action = 'unhidden'
    else:
        new_ignored = IgnoredDiscrepancy(user_id=current_user.id, hash_id=hash_id)
        db.session.add(new_ignored)
        action = 'hidden'
        
    db.session.commit()
    return jsonify({'success': True, 'action': action})

@report_bp.route('/global-schedule/history/unhide_all', methods=['POST'])
@login_required
def unhide_all_history():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.global_schedule_history'))
        
    from models import IgnoredDiscrepancy
    IgnoredDiscrepancy.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash('Se han vuelto a mostrar todos los registros.', 'success')
    return redirect(url_for('report.global_schedule_history'))

@report_bp.route('/global-schedule/export')
@login_required
def export_global_schedule():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    # Date Parameter
    date_str = request.args.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()

    # Time filter parameters
    time_from_str = request.args.get('time_from', '').strip()
    time_to_str = request.args.get('time_to', '').strip()
    time_from = None
    time_to = None
    try:
        if time_from_str:
            time_from = datetime.strptime(time_from_str, '%H:%M').time()
    except ValueError:
        pass
    try:
        if time_to_str:
            time_to = datetime.strptime(time_to_str, '%H:%M').time()
    except ValueError:
        pass

    # Query Matches
    matches = Match.query.join(League).filter(
        League.user_id == current_user.id,
        func.date(Match.match_date) == selected_date
    ).order_by(Match.match_date).all()

    # Apply time filter in Python
    if time_from or time_to:
        filtered = []
        for m in matches:
            t = m.match_date.time()
            if time_from and t < time_from:
                continue
            if time_to and t > time_to:
                continue
            filtered.append(m)
        matches = filtered
        
    # Inject Archived Matches
    archived = ArchivedFinance.query.filter(
        ArchivedFinance.user_id == current_user.id,
        ArchivedFinance.date == selected_date
    ).all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, selected_date.strftime('%Y-%m-%d'))
                    matches.append(fake_m)
            except:
                pass
                
    matches.sort(key=lambda x: x.match_date if x.match_date else datetime.max)

    # Group by Court
    grouped_schedule = {}
    for match in matches:
        court_name = match.court.name if match.court else "Sin Cancha Asignada"
        if court_name not in grouped_schedule:
            grouped_schedule[court_name] = []
        grouped_schedule[court_name].append(match)
    
    sorted_schedule = dict(sorted(grouped_schedule.items()))

    # Fetch Owner Court Settings
    owner_settings = {setting.court_name: setting.color for setting in OwnerCourtSetting.query.filter_by(user_id=current_user.id).all()}

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agenda Global"
    
    # Title
    time_range_label = ''
    if time_from_str or time_to_str:
        time_range_label = f" ({time_from_str or '00:00'} - {time_to_str or '23:59'})"
    ws['A1'] = f"AGENDA DE PARTIDOS - {selected_date.strftime('%d/%m/%Y')}{time_range_label}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    current_row = 3
    
    # Styles
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    
    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0
        try: return int(val)
        except: return 0

    for court_name, matches in sorted_schedule.items():
        # Court Header
        ws.cell(row=current_row, column=1, value=f"CANCHA: {court_name}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        
        court_color_hex = "2F855A" # Default green
        if court_name in owner_settings:
            raw_color = owner_settings[court_name]
            if raw_color and raw_color.startswith('#'):
                court_color_hex = raw_color[1:].upper()
                
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=court_color_hex, end_color=court_color_hex, fill_type="solid")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        current_row += 1
        
        # Table Headers
        headers = ["#", "Hora", "Categoría", "Local", "$ Local", "vs", "$ Visita", "Visitante", "$ Arbitro"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        total_home = 0
        total_away = 0
        total_referee = 0
        
        match_index = 1
        for match in matches:
            ws.cell(row=current_row, column=1, value=match_index).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=2, value=match.match_date.strftime('%I:%M %p'))
            
            league_cell = ws.cell(row=current_row, column=3, value=match.league.name)
            if match.league.custom_color_active and match.league.custom_name_color:
                league_color_hex = match.league.custom_name_color.lstrip('#').upper()
                league_cell.font = Font(bold=True, color=league_color_hex)
                
            ws.cell(row=current_row, column=4, value=match.home_team.name).alignment = Alignment(horizontal='right')
            
            c_home = parse_cost(match.referee_cost_home)
            ws.cell(row=current_row, column=5, value=c_home)
            total_home += c_home
            
            ws.cell(row=current_row, column=6, value="-").alignment = Alignment(horizontal='center')
            
            c_away = parse_cost(match.referee_cost_away)
            ws.cell(row=current_row, column=7, value=c_away)
            total_away += c_away
            
            ws.cell(row=current_row, column=8, value=match.away_team.name)
            
            c_ref = parse_cost(match.referee_cost)
            ws.cell(row=current_row, column=9, value=c_ref)
            total_referee += c_ref
            
            current_row += 1
            match_index += 1
            
        # Totals Row
        ws.cell(row=current_row, column=4, value="TOTALES:").alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=5, value=total_home).font = Font(bold=True)
        ws.cell(row=current_row, column=7, value=total_away).font = Font(bold=True)
        ws.cell(row=current_row, column=9, value=total_referee).font = Font(bold=True)
        current_row += 1
        
        # Teams Income
        teams_income = total_home + total_away
        ws.cell(row=current_row, column=8, value="TOTAL INGRESOS (Local + Visita):").alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=9, value=teams_income).font = Font(bold=True, color="0000FF") # Blue
        current_row += 1
        
        # Profit Row
        profit = teams_income - total_referee
        ws.cell(row=current_row, column=8, value="GANANCIA NETA (Ingresos - Árbitros):").alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=9, value=profit).font = Font(bold=True, color="008000" if profit >= 0 else "FF0000")
        current_row += 2 # Space between courts

    # Auto-adjust column widths
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Agenda_Global_{selected_date.strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)

# Helper to detect gifted/waived payments (any non-numeric, non-empty text like 'RG')
def is_waived(val):
    """Returns True if val is a non-numeric string indicating a gifted/waived payment."""
    if not val:
        return False
    try:
        int(str(val).strip())
        return False
    except (ValueError, TypeError):
        return True

# Helper for calculating discrepancies
def calculate_discrepancies(matches):
    events = []
    
    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0 
        try: return int(val)
        except: return 0

    for match in matches:
        if not match.league: continue
        
        # Respect Charge Start Date
        if not match.league.charge_from_start and match.league.charge_start_date:
            if match.match_date.date() < match.league.charge_start_date:
                continue

        # Defaults
        default_team_price = match.league.price_per_match or 0
        default_ref_price = match.league.price_referee or 0
        
        # Home – skip if waived (gifted)
        if not is_waived(match.referee_cost_home):
            paid_home = parse_cost(match.referee_cost_home)
            diff_home = paid_home - default_team_price
            if diff_home != 0:
                events.append({
                    'date': match.match_date,
                    'league': match.league.name,
                    'entity_type': 'Team',
                    'entity_name': match.home_team.name,
                    'balance': diff_home
                })
            
        # Away – skip if waived (gifted)
        if not is_waived(match.referee_cost_away):
            paid_away = parse_cost(match.referee_cost_away)
            diff_away = paid_away - default_team_price
            if diff_away != 0:
                events.append({
                    'date': match.match_date,
                    'league': match.league.name,
                    'entity_type': 'Team',
                    'entity_name': match.away_team.name,
                    'balance': diff_away
                })

        # Referee – skip if waived
        if not is_waived(match.referee_cost):
            paid_ref = parse_cost(match.referee_cost)
            diff_ref = paid_ref - default_ref_price
            if diff_ref != 0:
                events.append({
                    'date': match.match_date,
                    'league': match.league.name,
                    'entity_type': 'Referee',
                    'entity_name': 'Arbitro',
                    'balance': diff_ref
                })
    return events


@report_bp.route('/api/report/ignore_discrepancy', methods=['POST'])
@login_required
def ignore_discrepancy():
    if not getattr(current_user, 'is_ultra', False):
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    hash_id = data.get('hash_id')
    
    if not hash_id:
        return jsonify({'error': 'No info provided'}), 400
        
    # Check if already ignored to prevent duplicates
    existing = IgnoredDiscrepancy.query.filter_by(user_id=current_user.id, hash_id=hash_id).first()
    if not existing:
        new_ignore = IgnoredDiscrepancy(user_id=current_user.id, hash_id=hash_id)
        db.session.add(new_ignore)
        db.session.commit()
        
    return jsonify({'success': True})


def apply_multi_filters(query, archive_query, league_ids, cancha_names):
    if league_ids:
        query = query.filter(Match.league_id.in_(league_ids))
        league_objs = League.query.filter(League.id.in_(league_ids)).all()
        league_names = [l.name for l in league_objs]
        if league_names:
            archive_query = archive_query.filter(ArchivedFinance.league_name.in_(league_names))
            
    if cancha_names:
        if "Sin Cancha" in cancha_names:
            other_canchas = [c for c in cancha_names if c != "Sin Cancha"]
            if other_canchas:
                query = query.filter(db.or_(Match.court_id == None, Court.name.in_(other_canchas)))
                archive_query = archive_query.filter(db.or_(ArchivedFinance.court_name == "Sin Cancha", ArchivedFinance.court_name.in_(other_canchas)))
            else:
                query = query.filter(Match.court_id == None)
                archive_query = archive_query.filter(ArchivedFinance.court_name == "Sin Cancha")
        else:
            query = query.filter(Court.name.in_(cancha_names))
            archive_query = archive_query.filter(ArchivedFinance.court_name.in_(cancha_names))
            
    return query, archive_query

@report_bp.route('/global-schedule/summary')
@login_required
def global_schedule_summary():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    # Filters
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid] # remove empty strings
    
    cancha_names = request.args.getlist('cancha')
    cancha_names = [c for c in cancha_names if c]
    
    query = Match.query.join(League).outerjoin(Court, Match.court_id == Court.id).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, cancha_names)
            
    matches = query.all()
    archived = archive_query.all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, arc.date.strftime('%Y-%m-%d'))
                    fake_m.league.charge_from_start = True
                    fake_m.league.charge_start_date = None
                    matches.append(fake_m)
            except:
                pass
    
    discrepancies = calculate_discrepancies(matches)
    
    # Aggregate
    teams_summary = {} # (league, team_name) -> total
    referee_summary = {} # (league) -> total
    
    for item in discrepancies:
        if item['entity_type'] == 'Team':
            key = (item['league'], item['entity_name'])
            teams_summary[key] = teams_summary.get(key, 0) + item['balance']
            
            # The referee balance is now the sum of all team balances for that league
            league_key = item['league']
            referee_summary[league_key] = referee_summary.get(league_key, 0) + item['balance']
            
    # Convert to list for display, add hash ID
    teams_list = []
    for k, v in teams_summary.items():
        id_str = hashlib.md5(f"team_{k[0]}_{k[1]}".encode('utf-8')).hexdigest()
        teams_list.append({'league': k[0], 'name': k[1], 'balance': v, 'id': id_str})
    teams_list.sort(key=lambda x: (x['league'], x['name']))
    
    referee_list = []
    for k, v in referee_summary.items():
        id_str = hashlib.md5(f"ref_{k}".encode('utf-8')).hexdigest()
        referee_list.append({'league': k, 'balance': v, 'id': id_str})
    referee_list.sort(key=lambda x: x['league'])
    
    # Filter out excluded rows permanently from DB
    ignored_records = IgnoredDiscrepancy.query.filter_by(user_id=current_user.id).all()
    excluded_ids = [record.hash_id for record in ignored_records]
    
    if excluded_ids:
        teams_list = [t for t in teams_list if t['id'] not in excluded_ids]
        referee_list = [r for r in referee_list if r['id'] not in excluded_ids]
    
    # Get unique courts for dropdown
    courts_query = db.session.query(Court.name).join(League).filter(League.user_id == current_user.id).distinct().all()
    canchas = sorted([c[0] for c in courts_query if c[0]])
    
    leagues = League.query.filter_by(user_id=current_user.id).all()
    
    return render_template('report/summary.html', 
                         teams_summary=teams_list,
                         referee_summary=referee_list,
                         leagues=leagues, 
                         selected_league_ids=league_ids,
                         canchas=canchas,
                         selected_canchas=cancha_names)

@report_bp.route('/global-schedule/summary/share')
@login_required
def share_global_schedule_summary():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    # Filters
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid] # remove empty strings
    
    cancha_names = request.args.getlist('cancha')
    cancha_names = [c for c in cancha_names if c]
    
    query = Match.query.join(League).outerjoin(Court, Match.court_id == Court.id).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, cancha_names)
            
    matches = query.all()
    archived = archive_query.all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, arc.date.strftime('%Y-%m-%d'))
                    fake_m.league.charge_from_start = True
                    fake_m.league.charge_start_date = None
                    matches.append(fake_m)
            except:
                pass
    
    discrepancies = calculate_discrepancies(matches)
    
    # Aggregate
    teams_summary = {} # (league, team_name) -> total
    referee_summary = {} # (league) -> total
    
    for item in discrepancies:
        if item['entity_type'] == 'Team':
            key = (item['league'], item['entity_name'])
            teams_summary[key] = teams_summary.get(key, 0) + item['balance']
            
            # The referee balance is now the sum of all team balances for that league
            league_key = item['league']
            referee_summary[league_key] = referee_summary.get(league_key, 0) + item['balance']
            
    # Convert to list for display, add hash ID
    teams_list = []
    for k, v in teams_summary.items():
        id_str = hashlib.md5(f"team_{k[0]}_{k[1]}".encode('utf-8')).hexdigest()
        teams_list.append({'league': k[0], 'name': k[1], 'balance': v, 'id': id_str})
    teams_list.sort(key=lambda x: (x['league'], x['name']))
    
    referee_list = []
    for k, v in referee_summary.items():
        id_str = hashlib.md5(f"ref_{k}".encode('utf-8')).hexdigest()
        referee_list.append({'league': k, 'balance': v, 'id': id_str})
    referee_list.sort(key=lambda x: x['league'])
    
    # Filter out excluded rows permanently from DB
    ignored_records = IgnoredDiscrepancy.query.filter_by(user_id=current_user.id).all()
    excluded_ids = [record.hash_id for record in ignored_records]
    
    if excluded_ids:
        teams_list = [t for t in teams_list if t['id'] not in excluded_ids]
        referee_list = [r for r in referee_list if r['id'] not in excluded_ids]
    
    leagues = League.query.filter_by(user_id=current_user.id).all()
    
    if not league_ids:
        selected_league_name = 'Todas las Ligas'
    elif len(league_ids) == 1:
        selected_league_name = next((l.name for l in leagues if str(l.id) == league_ids[0]), 'Liga Seleccionada')
    else:
        selected_league_name = 'Varias Ligas'
        
    if not cancha_names:
        selected_cancha_display = 'Todas las Canchas'
    elif len(cancha_names) == 1:
        selected_cancha_display = cancha_names[0]
    else:
        selected_cancha_display = 'Varias Canchas'

    return render_template('report/share_summary.html', 
                         teams_summary=teams_list,
                         referee_summary=referee_list,
                         selected_league=selected_league_name,
                         selected_cancha=selected_cancha_display,
                         today=datetime.now().strftime('%d/%m/%Y'))

@report_bp.route('/global-schedule/summary/export')
@login_required
def export_global_summary():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    # Filters
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid] # remove empty strings
    
    cancha_names = request.args.getlist('cancha')
    cancha_names = [c for c in cancha_names if c]
    
    query = Match.query.join(League).outerjoin(Court, Match.court_id == Court.id).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, cancha_names)
            
    matches = query.all()
    archived = archive_query.all()
    import json
    for arc in archived:
        if arc.details_json:
            try:
                arc_data = json.loads(arc.details_json)
                for m_data in arc_data:
                    fake_m = FakeMatch(m_data, arc.league_name, arc.date.strftime('%Y-%m-%d'))
                    fake_m.league.charge_from_start = True
                    fake_m.league.charge_start_date = None
                    matches.append(fake_m)
            except:
                pass
                
    discrepancies = calculate_discrepancies(matches)
    
    # Aggregate
    teams_summary = {}
    referee_summary = {}
    
    for item in discrepancies:
        if item['entity_type'] == 'Team':
            key = (item['league'], item['entity_name'])
            teams_summary[key] = teams_summary.get(key, 0) + item['balance']
        elif item['entity_type'] == 'Referee':
            key = item['league']
            referee_summary[key] = referee_summary.get(key, 0) + item['balance']
            
    # Convert to list for display, add hash ID
    teams_list = []
    for k, v in teams_summary.items():
        id_str = hashlib.md5(f"team_{k[0]}_{k[1]}".encode('utf-8')).hexdigest()
        teams_list.append({'league': k[0], 'name': k[1], 'balance': v, 'id': id_str})
    teams_list.sort(key=lambda x: (x['league'], x['name']))
    
    referee_list = []
    for k, v in referee_summary.items():
        id_str = hashlib.md5(f"ref_{k}".encode('utf-8')).hexdigest()
        referee_list.append({'league': k, 'balance': v, 'id': id_str})
    referee_list.sort(key=lambda x: x['league'])
    
    # Filter out excluded rows permanently from DB
    ignored_records = IgnoredDiscrepancy.query.filter_by(user_id=current_user.id).all()
    excluded_ids = [record.hash_id for record in ignored_records]
    
    if excluded_ids:
        teams_list = [t for t in teams_list if t['id'] not in excluded_ids]
        referee_list = [r for r in referee_list if r['id'] not in excluded_ids]

    # Excel
    wb = openpyxl.Workbook()
    # Sheet 1: Teams
    ws1 = wb.active
    ws1.title = "Balance Equipos"
    ws1.append(["Liga", "Equipo", "Balance Total"])
    for t in teams_list:
        ws1.append([t['league'], t['name'], t['balance']])
        
    # Sheet 2: Referees
    ws2 = wb.create_sheet("Balance Arbitraje")
    ws2.append(["Liga", "Balance Total"])
    for r in referee_list:
        ws2.append([r['league'], r['balance']])
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Resumen_Global_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)

@report_bp.route('/global-schedule/financials')
@login_required
def global_schedule_financials():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    # Determine report type
    report_type = getattr(current_user, 'financial_report_type', 'period')

    # Obtener todas las canchas únicas de todas las ligas del owner
    leagues = current_user.leagues
    court_names = set()
    for league in leagues:
        for court in league.courts:
            if court.name:
                court_names.add(court.name.strip())
    court_names = sorted(list(court_names))

    cancha_names = request.args.getlist('cancha')
    cancha_names = [c for c in cancha_names if c]
    
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid]

    query = Match.query.join(League).outerjoin(Court, Match.court_id == Court.id).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, cancha_names)
    
    # Store applied filters to pass to template
    selected_month = None
    selected_year = None
    date_from_str = None
    date_to_str = None
    group_by = request.args.get('group_by', 'day')

    if report_type == 'date_range':
        # Handles date ranges
        default_from = datetime.now().replace(day=1).strftime('%Y-%m-%d')
        default_to = datetime.now().strftime('%Y-%m-%d')
        
        date_from_str = request.args.get('date_from', default=default_from)
        date_to_str = request.args.get('date_to', default=default_to)
        
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Match.match_date) >= d_from, func.date(Match.match_date) <= d_to)
        except ValueError:
            pass # fallback to no filter or default if desired, but here we just ignore invalid formats
            
    else:
        # Handles period (month/year)
        selected_month = request.args.get('month', type=int, default=datetime.now().month)
        selected_year = request.args.get('year', type=int, default=datetime.now().year)

        # Validate
        if not 1 <= selected_month <= 12: selected_month = datetime.now().month
        if selected_year < 2000 or selected_year > 2100: selected_year = datetime.now().year

        query = query.filter(
            func.extract('month', Match.match_date) == selected_month,
            func.extract('year', Match.match_date) == selected_year
        )

    matches = query.order_by(Match.match_date).all()

    # Structure: dict[date_str] = { 'date_obj': date, 'courts': { 'court_name': { income, expense, profit } }, 'daily_total': 0 }
    financial_data = {}
    
    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0 
        try: return int(val)
        except: return 0

    total_month_profit = 0
    court_totals = {} # Store total profit per court for the entire period

    for match in matches:
        if group_by == 'week':
            start_of_week = match.match_date - timedelta(days=match.match_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            date_key = f"Semana {start_of_week.isocalendar()[1]} ({start_of_week.strftime('%d/%m')} - {end_of_week.strftime('%d/%m')})"
            date_obj_sort = start_of_week
        elif group_by == 'month':
            months_es_dict = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
            month_name = months_es_dict.get(match.match_date.month, "")
            date_key = f"{month_name} {match.match_date.year}"
            date_obj_sort = match.match_date.replace(day=1)
        else: # day
            date_key = match.match_date.strftime('%d/%m/%Y')
            date_obj_sort = match.match_date

        court_name = match.court.name if match.court else "Sin Cancha"
        
        if not match.league: continue
        
        # Respect Charge Start Date
        if not match.league.charge_from_start and match.league.charge_start_date:
            if match.match_date.date() < match.league.charge_start_date:
                continue

        income = parse_cost(match.referee_cost_home) + parse_cost(match.referee_cost_away)
        expense = parse_cost(match.referee_cost)
        profit = income - expense
        
        if date_key not in financial_data:
            financial_data[date_key] = {
                'date_obj': date_obj_sort,
                'display_date': date_key,
                'courts': {},
                'daily_total': 0,
                'daily_income': 0,
                'daily_expense': 0
            }
            
        if court_name not in financial_data[date_key]['courts']:
            financial_data[date_key]['courts'][court_name] = {'income': 0, 'expense': 0, 'profit': 0}
            
        financial_data[date_key]['courts'][court_name]['income'] += income
        financial_data[date_key]['courts'][court_name]['expense'] += expense
        financial_data[date_key]['courts'][court_name]['profit'] += profit
        
        financial_data[date_key]['daily_total'] += profit
        financial_data[date_key]['daily_income'] += income
        financial_data[date_key]['daily_expense'] += expense
        total_month_profit += profit
        
        # Accumulate total profit per court
        if court_name not in court_totals:
            court_totals[court_name] = {'income': 0, 'expense': 0, 'profit': 0, 'dates': {}}
            
        court_totals[court_name]['income'] += income
        court_totals[court_name]['expense'] += expense
        court_totals[court_name]['profit'] += profit
        
        if date_key not in court_totals[court_name]['dates']:
            court_totals[court_name]['dates'][date_key] = {
                'date_obj': date_obj_sort,
                'display_date': date_key,
                'income': 0,
                'expense': 0,
                'profit': 0
            }
        
        court_totals[court_name]['dates'][date_key]['income'] += income
        court_totals[court_name]['dates'][date_key]['expense'] += expense
        court_totals[court_name]['dates'][date_key]['profit'] += profit

    # Process Archives
    
    if report_type == 'date_range' and date_from_str and date_to_str:
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            archive_query = archive_query.filter(ArchivedFinance.date >= d_from, ArchivedFinance.date <= d_to)
        except: pass
    elif selected_month and selected_year:
        archive_query = archive_query.filter(
            func.extract('month', ArchivedFinance.date) == selected_month,
            func.extract('year', ArchivedFinance.date) == selected_year
        )
        
    archived_finances = archive_query.all()
    total_month_profit = process_archives(archived_finances, group_by, financial_data, court_totals, total_month_profit)

    # Convert to sorted list
    sorted_data = sorted(financial_data.values(), key=lambda x: x['date_obj'])
    
    # Months for selector
    months_es = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"), (5, "Mayo"), (6, "Junio"),
        (7, "Julio"), (8, "Agosto"), (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre")
    ]
    
    # Years logic (current year -1 to +1)
    current_year = datetime.now().year
    years = [current_year - 1, current_year, current_year + 1]

    filename_suffix = ""
    if report_type == 'date_range':
        filename_suffix = f"{date_from_str}_al_{date_to_str}"
    else:
        month_name = dict(months_es).get(selected_month, selected_month)
        filename_suffix = f"{month_name}_{selected_year}"

    return render_template('report/financials.html', 
                         financial_data=sorted_data,
                         total_month_profit=total_month_profit,
                         selected_month=selected_month,
                         selected_year=selected_year,
                         date_from=date_from_str,
                         date_to=date_to_str,
                         report_type=report_type,
                         months=months_es,
                         years=years,
                         court_names=court_names,
                         selected_canchas=cancha_names,
                         leagues=leagues,
                         selected_league_ids=league_ids,
                         filename_suffix=filename_suffix,
                         court_totals=court_totals,
                         group_by=group_by)

@report_bp.route('/global-schedule/financials/export')
@login_required
def export_global_financials():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    report_type = getattr(current_user, 'financial_report_type', 'period')
    cancha_names = request.args.getlist('cancha')
    cancha_names = [c for c in cancha_names if c]
    
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid]
    
    group_by = request.args.get('group_by', 'day')
    query = Match.query.join(League).outerjoin(Court, Match.court_id == Court.id).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, cancha_names)
            
    filename_suffix = ""

    if report_type == 'date_range':
        date_from_str = request.args.get('date_from', default="")
        date_to_str = request.args.get('date_to', default="")
        
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Match.match_date) >= d_from, func.date(Match.match_date) <= d_to)
            filename_suffix = f"{date_from_str}_al_{date_to_str}"
        except ValueError:
            pass
            
    else:
        month = request.args.get('month', type=int, default=datetime.now().month)
        year = request.args.get('year', type=int, default=datetime.now().year)

        query = query.filter(
            func.extract('month', Match.match_date) == month,
            func.extract('year', Match.match_date) == year
        )
        filename_suffix = f"{month}_{year}"
        
    leagues = League.query.filter_by(user_id=current_user.id).all()
    if not league_ids:
        selected_league_name = 'Todas las Ligas'
    elif len(league_ids) == 1:
        selected_league_name = next((l.name for l in leagues if str(l.id) == league_ids[0]), 'Liga Seleccionada')
    else:
        selected_league_name = 'Varias Ligas'
    matches = query.order_by(Match.match_date).all()

    financial_data = {}
    court_totals = {}
    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0 
        try: return int(val)
        except: return 0

    for match in matches:
        if not match.league: continue
        
        # Respect Charge Start Date
        if not match.league.charge_from_start and match.league.charge_start_date:
            if match.match_date.date() < match.league.charge_start_date:
                continue

        if group_by == 'week':
            start_of_week = match.match_date - timedelta(days=match.match_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            date_key = f"Semana {start_of_week.isocalendar()[1]} ({start_of_week.strftime('%d/%m')} - {end_of_week.strftime('%d/%m')})"
            date_obj_sort = start_of_week
        elif group_by == 'month':
            months_es_dict = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
            month_name = months_es_dict.get(match.match_date.month, "")
            date_key = f"{month_name} {match.match_date.year}"
            date_obj_sort = match.match_date.replace(day=1)
        else: # day
            date_key = match.match_date.strftime('%d/%m/%Y')
            date_obj_sort = match.match_date

        court_name = match.court.name if match.court else "Sin Cancha"
        income = parse_cost(match.referee_cost_home) + parse_cost(match.referee_cost_away)
        expense = parse_cost(match.referee_cost)
        profit = income - expense
        
        if date_key not in financial_data:
            financial_data[date_key] = {'date_obj': date_obj_sort, 'display_date': date_key, 'courts': {}}
        if court_name not in financial_data[date_key]['courts']:
            financial_data[date_key]['courts'][court_name] = {'income': 0, 'expense': 0, 'profit': 0}
            
        financial_data[date_key]['courts'][court_name]['income'] += income
        financial_data[date_key]['courts'][court_name]['expense'] += expense
        financial_data[date_key]['courts'][court_name]['profit'] += profit
        
        # Accumulate total profit per court
        if court_name not in court_totals:
            court_totals[court_name] = {'income': 0, 'expense': 0, 'profit': 0, 'dates': {}}
            
        court_totals[court_name]['income'] += income
        court_totals[court_name]['expense'] += expense
        court_totals[court_name]['profit'] += profit
        
        if date_key not in court_totals[court_name]['dates']:
            court_totals[court_name]['dates'][date_key] = {
                'date_obj': date_obj_sort,
                'display_date': date_key,
                'income': 0,
                'expense': 0,
                'profit': 0
            }
        
        court_totals[court_name]['dates'][date_key]['income'] += income
        court_totals[court_name]['dates'][date_key]['expense'] += expense
        court_totals[court_name]['dates'][date_key]['profit'] += profit

    # Process Archives
    
    if report_type == 'date_range' and 'date_from_str' in locals() and 'date_to_str' in locals() and date_from_str:
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            archive_query = archive_query.filter(ArchivedFinance.date >= d_from, ArchivedFinance.date <= d_to)
        except: pass
    elif 'month' in locals() and 'year' in locals():
        archive_query = archive_query.filter(
            func.extract('month', ArchivedFinance.date) == month,
            func.extract('year', ArchivedFinance.date) == year
        )
        
    archived_finances = archive_query.all()
    process_archives(archived_finances, group_by, financial_data, court_totals)

    sorted_data = sorted(financial_data.values(), key=lambda x: x['date_obj'])
    
    leagues = League.query.filter_by(user_id=current_user.id).all()
    if not league_ids:
        selected_league_name = 'Todas las Ligas'
    elif len(league_ids) == 1:
        selected_league_name = next((l.name for l in leagues if str(l.id) == league_ids[0]), 'Liga Seleccionada')
    else:
        selected_league_name = 'Varias Ligas'

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Finanzas {filename_suffix}"[:31]  # Excel titles max 31 chars
    
    # Title
    header_title = filename_suffix.replace('_', ' ').upper()
    league_label = f" - {selected_league_name.upper()}" if selected_league_name else ""
    ws['A1'] = f"REPORTE FINANCIERO{league_label} - {header_title}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    current_row = 3
    headers = ["Fecha", "Cancha", "Ingresos", "Egresos", "Ganancia"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=h).font = Font(bold=True)
    current_row += 1
    
    total_profit = 0
    
    if report_type == 'por_cancha':
        for court_name, c_stats in court_totals.items():
            # Court Header Row
            ws.cell(row=current_row, column=1, value=court_name.upper())
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="008000")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            current_row += 1
            
            # Sorted inner dates
            sorted_dates = sorted(c_stats['dates'].values(), key=lambda x: x['date_obj'])
            
            for d_stats in sorted_dates:
                date_str = d_stats['display_date']
                
                ws.cell(row=current_row, column=1, value=date_str)
                ws.cell(row=current_row, column=2, value=court_name)
                ws.cell(row=current_row, column=3, value=d_stats['income'])
                ws.cell(row=current_row, column=4, value=d_stats['expense'])
                ws.cell(row=current_row, column=5, value=d_stats['profit'])
                
                # Color profit
                color = "008000" if d_stats['profit'] >= 0 else "FF0000"
                ws.cell(row=current_row, column=5).font = Font(color=color, bold=True)
                current_row += 1
                
            # Court Subtotal Row
            ws.cell(row=current_row, column=2, value="SUBTOTAL CANCHA:").alignment = Alignment(horizontal='right')
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=c_stats['income']).font = Font(color="0000FF", bold=True)
            ws.cell(row=current_row, column=3).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            ws.cell(row=current_row, column=4, value=c_stats['expense']).font = Font(color="FF0000", bold=True)
            ws.cell(row=current_row, column=4).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            ws.cell(row=current_row, column=5, value=c_stats['profit']).font = Font(bold=True)
            ws.cell(row=current_row, column=5).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            current_row += 1
            total_profit += c_stats['profit']
            
    else:
        for day in sorted_data:
            date_str = day['display_date']
            daily_profit = 0
            daily_income = 0
            daily_expense = 0
            
            for court_name, stats in day['courts'].items():
                ws.cell(row=current_row, column=1, value=date_str)
                ws.cell(row=current_row, column=2, value=court_name)
                ws.cell(row=current_row, column=3, value=stats['income'])
                ws.cell(row=current_row, column=4, value=stats['expense'])
                ws.cell(row=current_row, column=5, value=stats['profit'])
                
                # Color profit
                color = "008000" if stats['profit'] >= 0 else "FF0000"
                ws.cell(row=current_row, column=5).font = Font(color=color, bold=True)
                
                daily_profit += stats['profit']
                daily_income += stats['income']
                daily_expense += stats['expense']
                current_row += 1
                
            # Daily Total Row
            ws.cell(row=current_row, column=2, value="TOTAL DÍA:").alignment = Alignment(horizontal='right')
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=daily_income).font = Font(bold=True)
            ws.cell(row=current_row, column=3).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            ws.cell(row=current_row, column=4, value=daily_expense).font = Font(color="FF0000", bold=True)
            ws.cell(row=current_row, column=4).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            ws.cell(row=current_row, column=5, value=daily_profit).font = Font(bold=True)
            ws.cell(row=current_row, column=5).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            current_row += 1
            total_profit += daily_profit
        
    current_row += 1
    ws.cell(row=current_row, column=4, value="GRAN TOTAL PERIODO:").alignment = Alignment(horizontal='right')
    ws.cell(row=current_row, column=5, value=total_profit).font = Font(bold=True, size=12)
    
    # Adjust widths
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Finanzas_{filename_suffix}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)

@report_bp.route('/global-schedule/financials/share')
@login_required
def share_global_financials():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))

    report_type = getattr(current_user, 'financial_report_type', 'period')
    cancha_names = request.args.getlist('cancha')
    cancha_names = [c for c in cancha_names if c]
    
    league_ids = request.args.getlist('league_id')
    league_ids = [lid for lid in league_ids if lid]
    
    group_by = request.args.get('group_by', 'day')
    query = Match.query.join(League).outerjoin(Court, Match.court_id == Court.id).filter(League.user_id == current_user.id)
    archive_query = ArchivedFinance.query.filter_by(user_id=current_user.id)
    
    query, archive_query = apply_multi_filters(query, archive_query, league_ids, cancha_names)
            
    header_title = ""

    if report_type == 'date_range':
        date_from_str = request.args.get('date_from', default="")
        date_to_str = request.args.get('date_to', default="")
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Match.match_date) >= d_from, func.date(Match.match_date) <= d_to)
            header_title = f"DESDE {date_from_str} AL {date_to_str}"
        except ValueError:
            header_title = "RANGO DE FECHAS"
            pass
    else:
        month = request.args.get('month', type=int, default=datetime.now().month)
        year = request.args.get('year', type=int, default=datetime.now().year)

        query = query.filter(
            func.extract('month', Match.match_date) == month,
            func.extract('year', Match.match_date) == year
        )
        months_es = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
        month_name = months_es.get(month, "")
        header_title = f"{month_name} {year}"
        
    matches = query.order_by(Match.match_date).all()

    financial_data = {}
    total_income = 0
    total_expense = 0
    total_profit = 0
    court_totals = {}

    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0 
        try: return int(val)
        except: return 0

    for match in matches:
        if not match.league: continue
        
        # Respect Charge Start Date
        if not match.league.charge_from_start and match.league.charge_start_date:
            if match.match_date.date() < match.league.charge_start_date:
                continue

        if group_by == 'week':
            start_of_week = match.match_date - timedelta(days=match.match_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            date_key = f"Semana {start_of_week.isocalendar()[1]} ({start_of_week.strftime('%d/%m')} - {end_of_week.strftime('%d/%m')})"
            date_obj_sort = start_of_week
        elif group_by == 'month':
            months_es_dict = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
            month_name = months_es_dict.get(match.match_date.month, "")
            date_key = f"{month_name} {match.match_date.year}"
            date_obj_sort = match.match_date.replace(day=1)
        else: # day
            date_key = match.match_date.strftime('%d/%m/%Y')
            date_obj_sort = match.match_date

        court_name = match.court.name if match.court else "Sin Cancha"
        income = parse_cost(match.referee_cost_home) + parse_cost(match.referee_cost_away)
        expense = parse_cost(match.referee_cost)
        profit = income - expense
        
        if date_key not in financial_data:
            financial_data[date_key] = {'date_obj': date_obj_sort, 'display_date': date_key, 'courts': {}, 'daily_income': 0, 'daily_expense': 0, 'daily_profit': 0}
        if court_name not in financial_data[date_key]['courts']:
            financial_data[date_key]['courts'][court_name] = {'income': 0, 'expense': 0, 'profit': 0}
            
        financial_data[date_key]['courts'][court_name]['income'] += income
        financial_data[date_key]['courts'][court_name]['expense'] += expense
        financial_data[date_key]['courts'][court_name]['profit'] += profit
        
        financial_data[date_key]['daily_income'] += income
        financial_data[date_key]['daily_expense'] += expense
        financial_data[date_key]['daily_profit'] += profit
        
        # Accumulate total profit per court
        if court_name not in court_totals:
            court_totals[court_name] = {'income': 0, 'expense': 0, 'profit': 0, 'dates': {}}
            
        court_totals[court_name]['income'] += income
        court_totals[court_name]['expense'] += expense
        court_totals[court_name]['profit'] += profit
        
        if date_key not in court_totals[court_name]['dates']:
            court_totals[court_name]['dates'][date_key] = {
                'date_obj': date_obj_sort,
                'display_date': date_key,
                'income': 0,
                'expense': 0,
                'profit': 0
            }
        
        court_totals[court_name]['dates'][date_key]['income'] += income
        court_totals[court_name]['dates'][date_key]['expense'] += expense
        court_totals[court_name]['dates'][date_key]['profit'] += profit

        total_income += income
        total_expense += expense
        total_profit += profit

    # Process Archives
    
    if report_type == 'date_range' and 'date_from_str' in locals() and 'date_to_str' in locals() and date_from_str:
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            archive_query = archive_query.filter(ArchivedFinance.date >= d_from, ArchivedFinance.date <= d_to)
        except: pass
    elif 'month' in locals() and 'year' in locals():
        archive_query = archive_query.filter(
            func.extract('month', ArchivedFinance.date) == month,
            func.extract('year', ArchivedFinance.date) == year
        )
        
    archived_finances = archive_query.all()
    # share_global_financials also keeps track of global total_income, total_expense, total_profit
    for archive in archived_finances:
        total_income += archive.income
        total_expense += archive.expense
        total_profit += archive.profit
        
    process_archives(archived_finances, group_by, financial_data, court_totals)

    sorted_data = sorted(financial_data.values(), key=lambda x: x['date_obj'])

    leagues = League.query.filter_by(user_id=current_user.id).all()
    if not league_ids:
        selected_league_name = 'Todas las Ligas'
    elif len(league_ids) == 1:
        selected_league_name = next((l.name for l in leagues if str(l.id) == league_ids[0]), 'Liga Seleccionada')
    else:
        selected_league_name = 'Varias Ligas'

    if not cancha_names:
        selected_cancha_name = 'TODAS LAS CANCHAS'
    elif len(cancha_names) == 1:
        selected_cancha_name = cancha_names[0].upper()
    else:
        selected_cancha_name = 'VARIAS CANCHAS'

    return render_template('report/share_financials.html', 
                         financial_data=sorted_data,
                         header_title=header_title,
                         selected_cancha_name=selected_cancha_name,
                         selected_league_name=selected_league_name,
                         total_income=total_income,
                         total_expense=total_expense,
                         total_profit=total_profit,
                         today=datetime.now().strftime('%d/%m/%Y'),
                         report_type=report_type,
                         court_totals=court_totals,
                         group_by=group_by)

@report_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if not getattr(current_user, 'is_ultra', False):
        flash('No tienes acceso a esta funcionalidad (Ultra Premium).', 'warning')
        return redirect(url_for('report.index'))

    # Obtener todas las canchas únicas de todas las ligas del owner
    leagues = current_user.leagues
    court_names = set()
    for league in leagues:
        for court in league.courts:
            if court.name:
                court_names.add(court.name.strip())

    court_names = sorted(list(court_names))

    if request.method == 'POST':
        # Guardar configuraciones de canchas
        for court_name in court_names:
            color = request.form.get(f'color_{court_name}')
            if color:
                setting = OwnerCourtSetting.query.filter_by(user_id=current_user.id, court_name=court_name).first()
                if not setting:
                    setting = OwnerCourtSetting(user_id=current_user.id, court_name=court_name)
                    db.session.add(setting)
                setting.color = color
                
        # Guardar configuración de tipo de reporte financiero
        report_type = request.form.get('financial_report_type')
        if report_type in ['period', 'date_range', 'por_cancha']:
            current_user.financial_report_type = report_type

        
        db.session.commit()
        flash('Configuraciones generales de reportes guardadas con éxito.', 'success')
        return redirect(url_for('report.index'))

    # Cargar configuraciones existentes
    existing_settings = OwnerCourtSetting.query.filter_by(user_id=current_user.id).all()
    court_colors = {s.court_name: s.color for s in existing_settings}

    return render_template('report/settings.html', court_names=court_names, court_colors=court_colors)

@report_bp.route('/financial-report/charts')
@login_required
def financial_charts():
    if not getattr(current_user, 'is_ultra', False):
        flash('No tienes acceso a esta funcionalidad (Ultra Premium).', 'warning')
        return redirect(url_for('report.index'))
    
    # Get all unique court names for filtering
    leagues = League.query.filter_by(user_id=current_user.id).all()
    court_names = set()
    for l in leagues:
        for c in l.courts:
            if c.name:
                court_names.add(c.name.strip())
    
    sorted_court_names = sorted(list(court_names))
    
    return render_template('report/financial_charts.html', court_names=sorted_court_names)

@report_bp.route('/api/report/financial-stats')
@login_required
def api_financial_stats():
    if not getattr(current_user, 'is_ultra', False):
        return jsonify({'error': 'Unauthorized'}), 401
        
    period = request.args.get('period', 'day') # day, week, month
    court_name = request.args.get('court_name') # optional
    date_from_str = request.args.get('date_from') # optional (YYYY-MM-DD)
    date_to_str = request.args.get('date_to') # optional (YYYY-MM-DD)
    
    # Base query: matches from leagues owned by current_user
    query = Match.query.join(League).filter(League.user_id == current_user.id)
    
    if court_name:
        query = query.outerjoin(Court, Match.court_id == Court.id).filter(Court.name == court_name)
    
    if date_from_str:
        try:
            d_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Match.match_date) >= d_from)
        except ValueError:
            pass
            
    if date_to_str:
        try:
            d_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(func.date(Match.match_date) <= d_to)
        except ValueError:
            pass
        
    matches = query.order_by(Match.match_date).all()
    
    stats_data = {} # label -> profit
    
    def parse_cost(val):
        if not val: return 0
        if isinstance(val, str) and not val.isdigit(): return 0 
        try: return int(val)
        except: return 0

    for match in matches:
        if not match.match_date: continue
        
        # Respect Charge Start Date
        if not match.league.charge_from_start and match.league.charge_start_date:
            if match.match_date.date() < match.league.charge_start_date:
                continue
                
        # Grouping label
        if period == 'day':
            label = match.match_date.strftime('%Y-%m-%d')
        elif period == 'week':
            # ISO format for sorting: YYYY-Www
            sort_key = match.match_date.strftime('%G-W%V')
            # Calculate Monday of the week for display
            monday = match.match_date - timedelta(days=match.match_date.weekday())
            # Format: "Sem. 11 (16/03)"
            display_label = f"Sem. {match.match_date.strftime('%V')} ({monday.strftime('%d/%m')})"
            label = (sort_key, display_label)
        elif period == 'month':
            label = match.match_date.strftime('%Y-%m')
        else:
            label = match.match_date.strftime('%Y-%m-%d')
            
        profit = (parse_cost(match.referee_cost_home) + parse_cost(match.referee_cost_away)) - parse_cost(match.referee_cost)
        
        stats_data[label] = stats_data.get(label, 0) + profit
        
    # Sort byproduct by label
    sorted_keys = sorted(stats_data.keys())
    
    # Extract final labels and values
    final_labels = [k[1] if isinstance(k, tuple) else k for k in sorted_keys]
    values = [stats_data[k] for k in sorted_keys]
    
    return jsonify({
        'labels': final_labels,
        'values': values
    })

@report_bp.route('/archived-finances')
@report_bp.route('/archived-finances')
@login_required
def archived_finances():
    if not getattr(current_user, 'is_ultra', False):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('report.index'))
        
    from models import ArchivedFinance
    import json
    
    # Sort by created_at desc so newest batches appear first
    archives = ArchivedFinance.query.filter_by(user_id=current_user.id).order_by(ArchivedFinance.created_at.desc(), ArchivedFinance.date.desc()).all()
    
    batches_dict = {}
    
    for arc in archives:
        # Group by league_name and created_at (formatted to minute to catch items created in the same loop)
        batch_key = f"{arc.league_name}_{arc.created_at.strftime('%Y%m%d%H%M') if arc.created_at else arc.date.strftime('%Y%m%d')}"
        
        if batch_key not in batches_dict:
            batches_dict[batch_key] = {
                'id': batch_key,
                'league_name': arc.league_name,
                'archived_date': arc.created_at if arc.created_at else arc.date,
                'courts': set(),
                'income': 0,
                'expense': 0,
                'profit': 0,
                'archive_ids': [],
                'matches': []
            }
            
        b = batches_dict[batch_key]
        b['courts'].add(arc.court_name)
        b['income'] += arc.income
        b['expense'] += arc.expense
        b['profit'] += arc.profit
        b['archive_ids'].append(arc.id)
        
        if arc.details_json:
            try:
                matches_list = json.loads(arc.details_json)
                for idx, m in enumerate(matches_list):
                    m['arc_id'] = arc.id
                    m['match_idx'] = idx
                    m['court_name'] = arc.court_name
                    m['match_date'] = arc.date
                    b['matches'].append(m)
            except:
                pass

    batch_list = list(batches_dict.values())
    batch_list.sort(key=lambda x: x['archived_date'], reverse=True)
    
    for b in batch_list:
        b['courts'] = ", ".join(sorted(list(b['courts'])))
        b['archive_ids_json'] = json.dumps(b['archive_ids'])
        
    return render_template('report/archived_finances.html', batches=batch_list)

@report_bp.route('/api/archived-finances/batch', methods=['DELETE'])
@login_required
def delete_archived_finance_batch():
    if not getattr(current_user, 'is_ultra', False):
        return jsonify({'success': False, 'message': 'Acceso denegado.'}), 403
        
    from models import ArchivedFinance
    data = request.get_json() or {}
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'success': False, 'message': 'No se proporcionaron IDs.'}), 400
        
    for arc_id in ids:
        archive = ArchivedFinance.query.filter_by(id=arc_id, user_id=current_user.id).first()
        if archive:
            db.session.delete(archive)
            
    db.session.commit()
    return jsonify({'success': True, 'message': 'Lote eliminado correctamente.'})

@report_bp.route('/api/archived-finances/<arc_id>/match/<int:match_idx>', methods=['DELETE'])
@login_required
def delete_archived_finance_match(arc_id, match_idx):
    if not getattr(current_user, 'is_ultra', False):
        return jsonify({'success': False, 'message': 'Acceso denegado.'}), 403
        
    from models import ArchivedFinance
    import json
    
    archive = ArchivedFinance.query.filter_by(id=arc_id, user_id=current_user.id).first()
    if not archive:
        return jsonify({'success': False, 'message': 'Registro no encontrado.'}), 404
        
    if not archive.details_json:
        return jsonify({'success': False, 'message': 'El registro no tiene desglose de partidos.'}), 400
        
    try:
        matches = json.loads(archive.details_json)
        if match_idx < 0 or match_idx >= len(matches):
            return jsonify({'success': False, 'message': 'Índice de partido inválido.'}), 400
            
        match = matches.pop(match_idx)
        
        archive.income -= (int(match.get('home_paid', 0)) + int(match.get('away_paid', 0)))
        archive.expense -= int(match.get('ref_paid', 0))
        archive.profit = archive.income - archive.expense
        
        if archive.income == 0 and archive.expense == 0:
            db.session.delete(archive)
        else:
            archive.details_json = json.dumps(matches)
            
        db.session.commit()
        return jsonify({'success': True, 'message': 'Partido eliminado correctamente.'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
